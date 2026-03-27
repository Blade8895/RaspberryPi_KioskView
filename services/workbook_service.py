from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


class WorkbookService:
    def __init__(self, logger):
        self.logger = logger

    @staticmethod
    def _fmt(v):
        if v is None:
            return ""
        if isinstance(v, (date, datetime)):
            return v.strftime("%d.%m.%Y")
        return str(v)

    @staticmethod
    def _color_to_css(color):
        if not color:
            return None
        color_type = getattr(color, "type", None)
        if color_type == "rgb":
            rgb = getattr(color, "rgb", None)
            if not rgb:
                return None
            rgb = rgb.upper()
            if len(rgb) == 8:  # ARGB -> RGB
                rgb = rgb[2:]
            if len(rgb) == 6:
                return f"#{rgb}"
        return None

    def _extract_style(self, cell):
        font = cell.font
        fill = cell.fill
        alignment = cell.alignment
        border = cell.border

        style = {
            "font": {
                "bold": bool(font.bold),
                "italic": bool(font.italic),
                "name": font.name,
                "size": font.sz,
                "color": self._color_to_css(font.color),
            },
            "fill": {
                "background": self._color_to_css(getattr(fill, "fgColor", None)),
            },
            "alignment": {
                "horizontal": alignment.horizontal,
                "vertical": alignment.vertical,
                "wrap_text": bool(alignment.wrap_text),
            },
            "border": {
                "left": {
                    "style": border.left.style,
                    "color": self._color_to_css(border.left.color),
                },
                "right": {
                    "style": border.right.style,
                    "color": self._color_to_css(border.right.color),
                },
                "top": {
                    "style": border.top.style,
                    "color": self._color_to_css(border.top.color),
                },
                "bottom": {
                    "style": border.bottom.style,
                    "color": self._color_to_css(border.bottom.color),
                },
            },
        }
        return style

    def _build_layout(self, ws, min_col, min_row, max_col, max_row):
        row_heights = []
        for r in range(min_row, max_row + 1):
            row_heights.append(ws.row_dimensions[r].height)

        column_widths = []
        for c in range(min_col, max_col + 1):
            letter = get_column_letter(c)
            column_widths.append(ws.column_dimensions[letter].width)

        merges = []
        merge_map = {}
        for merged in ws.merged_cells.ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))

            i_min_col = max(min_col, m_min_col)
            i_min_row = max(min_row, m_min_row)
            i_max_col = min(max_col, m_max_col)
            i_max_row = min(max_row, m_max_row)
            if i_min_col > i_max_col or i_min_row > i_max_row:
                continue

            row_span = i_max_row - i_min_row + 1
            col_span = i_max_col - i_min_col + 1
            merges.append(
                {
                    "row": i_min_row - min_row,
                    "col": i_min_col - min_col,
                    "row_span": row_span,
                    "col_span": col_span,
                    "source_range": str(merged),
                }
            )

            for r in range(i_min_row, i_max_row + 1):
                for c in range(i_min_col, i_max_col + 1):
                    key = (r - min_row, c - min_col)
                    merge_map[key] = {
                        "row": i_min_row - min_row,
                        "col": i_min_col - min_col,
                        "row_span": row_span,
                        "col_span": col_span,
                        "hidden": not (r == i_min_row and c == i_min_col),
                    }

        layout = {
            "row_heights": row_heights,
            "column_widths": column_widths,
            "merged_ranges": merges,
        }
        return layout, merge_map

    def read_view(self, xlsx_path: str, sheet_name: str, cell_range: str):
        wb = load_workbook(xlsx_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' nicht gefunden")
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        layout, merge_map = self._build_layout(ws, min_col, min_row, max_col, max_row)

        cells = []
        rows = []
        for r in range(min_row, max_row + 1):
            row_values = []
            row_cells = []
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                value = self._fmt(cell.value)
                row_values.append(value)

                rel_key = (r - min_row, c - min_col)
                row_cells.append(
                    {
                        "value": value,
                        "style": self._extract_style(cell),
                        "merge": merge_map.get(rel_key),
                    }
                )
            rows.append(row_values)
            cells.append(row_cells)

        return {
            "sheet": sheet_name,
            "range": cell_range,
            "rows": rows,
            "cells": cells,
            "layout": layout,
        }
